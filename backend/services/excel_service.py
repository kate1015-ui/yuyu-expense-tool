"""填寫「出差旅費報告表」Excel 模板並回傳二進位檔案。

表格欄位對應 (依空白表單反推):
  C4  姓名       E4  單位(嶼嶼行銷)  I4  職稱
  C5  事由
  C6  期間起日   G6  期間迄日
  --- 明細列 R9-R14 (最多 6 列) ---
  A   日期       C   起訖地點
  D   交通工具   E   交通費金額
  H   膳雜費金額 (僅第一列填入, 後列空白)
  I   摘要 (代墊時填 &同行人)
  --- 合計 ---
  I15 人數       I16 總合計
  E16 交通費合計 H16 膳雜費合計
"""
from __future__ import annotations

import copy
import os
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "expense_template.xlsx")
TEMPLATE_SHEET = "空白表單"
MEAL_PER_PERSON = 700   # 每人膳雜費 (NTD)
DATA_ROW_START = 9
DATA_ROW_END   = 14


def _fmt_date(d) -> Optional[datetime]:
    """接受 str(YYYY-MM-DD) 或 datetime, 回傳 datetime 供 openpyxl 寫入。"""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day)
    try:
        return datetime.strptime(str(d).strip(), "%Y-%m-%d")
    except ValueError:
        return None


def generate_expense_excel(payload: dict) -> bytes:
    """填入 payload 並回傳 .xlsx 二進位。

    payload 必填:
      name        str   姓名
      title       str   職稱
      reason      str   事由  (例: 拍攝珮瑜)
      date_from   str   期間起 YYYY-MM-DD
      date_to     str   期間迄 YYYY-MM-DD
      people      int   人數 (影響膳雜費計算)
      rows        list  明細列 (見下)
        - date          str   日期
        - route         str   起訖地點 (例: 公司-珮瑜公司\n(19.6公里))
        - transport     str   交通工具 (自行開車 / 大眾交通)
        - transport_amt float 交通費金額
        - meal_amt      float 膳雜費 (只有第一列填, 其餘 0)
        - note          str   摘要 (代墊時填 &同行人姓名)
    選填:
      unit        str   單位 (預設 嶼嶼行銷)
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise RuntimeError(f"找不到 Excel 模板: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    # 以空白表單為基礎複製一份新工作表
    src = wb[TEMPLATE_SHEET]
    new_title = f"報帳_{payload.get('reason','')[:8]}"[:31]
    # openpyxl copy_worksheet
    dest = wb.copy_worksheet(src)
    dest.title = new_title

    # ---- 標頭 ----
    dest["C4"] = payload.get("name", "")
    dest["E4"] = payload.get("unit", "嶼嶼行銷")
    dest["I4"] = payload.get("title", "")
    dest["C5"] = payload.get("reason", "")
    dest["C6"] = _fmt_date(payload.get("date_from"))
    dest["G6"] = _fmt_date(payload.get("date_to"))

    # ---- 明細列 ----
    rows = payload.get("rows", [])
    total_transport = 0.0
    total_meal = 0.0

    for idx in range(DATA_ROW_END - DATA_ROW_START + 1):
        r = DATA_ROW_START + idx
        if idx < len(rows):
            row_data = rows[idx]
            dest[f"A{r}"] = _fmt_date(row_data.get("date"))
            dest[f"C{r}"] = row_data.get("route", "")
            dest[f"D{r}"] = row_data.get("transport", "")
            t_amt = float(row_data.get("transport_amt") or 0)
            dest[f"E{r}"] = t_amt
            total_transport += t_amt
            m_amt = float(row_data.get("meal_amt") or 0)
            if m_amt:
                dest[f"H{r}"] = m_amt
                total_meal += m_amt
            dest[f"I{r}"] = row_data.get("note", "")
        else:
            # 清空多餘列
            for col in ("A", "C", "D", "E", "F", "G", "H", "I"):
                dest[f"{col}{r}"] = None

    # ---- 人數 & 合計 ----
    people = int(payload.get("people", 1))
    dest["I15"] = people
    dest["E16"] = round(total_transport, 1)
    dest["G16"] = 0
    dest["H16"] = round(total_meal, 0)
    dest["I16"] = round(total_transport + total_meal, 1)

    # ---- 輸出 ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_rows_from_form(form: dict) -> list[dict]:
    """前端 payload → 明細列 list。

    form 欄位:
      transport_mode  'drive' | 'transit'
      date            str
      legs            list  (新版格式，見下)
      people          int
      companions      str
      pay_for_others  bool

    transit legs 每項:
      {origin, destination, tool, amount}
      tool: "計程車" | "高鐵" | "台鐵" | "捷運"

    drive legs 每項:
      {description, origin, destination, distance_km, cost}
    """
    d = form.get("date", "")
    mode = form.get("transport_mode")
    companions = form.get("companions", "").strip()

    note = ""
    if companions:
        names = [n.strip() for n in companions.replace("、", ",").split(",") if n.strip()]
        if names:
            note = "幫" + "、".join(names) + "代墊"

    people = int(form.get("people", 1))
    meal_total = MEAL_PER_PERSON * people
    legs = form.get("legs", [])

    rows = []
    first = True

    if mode == "drive":
        parking_total = 0.0
        for leg in legs:
            cost = float(leg.get("cost") or 0)
            parking_total += float(leg.get("parking") or 0)
            desc = leg.get("description", "").strip()
            distance = leg.get("distance_km")
            route_label = desc if desc else f"{leg.get('origin','')}→{leg.get('destination','')}"
            if distance is not None:
                route_label = f"{route_label}\n(共{distance}公里)"
            rows.append({
                "date": d,
                "route": route_label,
                "transport": "自行開車",
                "transport_amt": cost,
                "meal_amt": meal_total if first else 0,
                "note": note if first else "",
            })
            first = False
        # 停車費：獨立一列
        if parking_total > 0:
            rows.append({
                "date": d,
                "route": "停車費",
                "transport": "停車費",
                "transport_amt": parking_total,
                "meal_amt": 0,
                "note": "",
            })
        # ETag / 過路費：獨立一列
        etag_total = float(form.get("etag_amt") or 0)
        if etag_total > 0:
            rows.append({
                "date": d,
                "route": "ETag/過路費",
                "transport": "ETag/過路費",
                "transport_amt": etag_total,
                "meal_amt": 0,
                "note": "",
            })
    elif mode == "transit":
        for leg in legs:
            amt = float(leg.get("amount") or 0)
            origin = leg.get("origin", "").strip()
            dest = leg.get("destination", "").strip()
            tool = (leg.get("tool") or "大眾交通").strip()
            rows.append({
                "date": d,
                "route": f"{origin}→{dest}" if origin or dest else "",
                "transport": tool,
                "transport_amt": amt,
                "meal_amt": meal_total if first else 0,
                "note": note if first else "",
            })
            first = False

    return rows
